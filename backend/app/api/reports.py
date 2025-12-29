"""Reports API routes - Read-only GST reports for internal review

IMPORTANT:
- All reports are for INTERNAL BUSINESS REVIEW only
- NOT for GST return filing
- GST filing must be done on official GST portal
"""

from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import csv
import io
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.api.auth import get_current_active_user
from app.models.user import User
from app.schemas.report import (
    SalesRegisterResponse,
    PurchaseRegisterResponse,
    GSTSummaryResponse,
    CustomerReportResponse,
    SupplierReportResponse,
    ProductHSNReportResponse,
    InventoryReportResponse,
    BusinessSummaryLedger,
    GSTRReadyResponse
)
from app.services import report_service

router = APIRouter()


@router.get("/sales", response_model=SalesRegisterResponse)
def get_sales_register(
    from_date: date = Query(..., description="Start date (inclusive)"),
    to_date: date = Query(..., description="End date (inclusive)"),
    include_cancelled: bool = Query(False, description="Include cancelled invoices in totals"),
    customer_id: Optional[int] = Query(None, description="Filter by customer ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate Sales Register report.
    
    Lists all sales invoices for the given date range with GST breakdown.
    Cancelled invoices appear in the list but are excluded from totals (unless include_cancelled=True).
    
    NOTE: This report is for internal review only. NOT for GST filing.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_sales_register(
        db=db,
        from_date=from_date,
        to_date=to_date,
        include_cancelled=include_cancelled,
        customer_id=customer_id
    )


@router.get("/purchases", response_model=PurchaseRegisterResponse)
def get_purchase_register(
    from_date: date = Query(..., description="Start date (inclusive)"),
    to_date: date = Query(..., description="End date (inclusive)"),
    include_cancelled: bool = Query(False, description="Include cancelled purchases in totals"),
    supplier_id: Optional[int] = Query(None, description="Filter by supplier ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate Purchase Register report.
    
    Lists all purchase invoices for the given date range with input GST breakdown.
    Cancelled purchases appear in the list but are excluded from totals (unless include_cancelled=True).
    
    NOTE: This report is for internal review only. NOT for ITC claims.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_purchase_register(
        db=db,
        from_date=from_date,
        to_date=to_date,
        include_cancelled=include_cancelled,
        supplier_id=supplier_id
    )


@router.get("/gst-summary", response_model=GSTSummaryResponse)
def get_gst_summary(
    month: int = Query(..., ge=1, le=12, description="Month (1-12)"),
    year: int = Query(..., ge=2000, le=2100, description="Year"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate monthly GST summary.
    
    Provides high-level view of:
    - Output GST (from sales)
    - Input GST (from purchases)
    
    NOTE: This report does NOT calculate payable GST or ITC offset.
    For GST filing, use the official GST portal.
    """
    return report_service.generate_gst_summary(
        db=db,
        month=month,
        year=year
    )


@router.get("/sales/export/csv")
def export_sales_register_csv(
    from_date: date = Query(...),
    to_date: date = Query(...),
    include_cancelled: bool = Query(False),
    customer_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Sales Register to CSV.
    
    CSV format matches on-screen report exactly.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    report = report_service.generate_sales_register(
        db=db,
        from_date=from_date,
        to_date=to_date,
        include_cancelled=include_cancelled,
        customer_id=customer_id
    )
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Invoice Number', 'Invoice Date', 'Customer Name', 'Customer GSTIN',
        'Place of Supply', 'Taxable Value', 'CGST', 'SGST', 'IGST',
        'Total GST', 'Grand Total', 'Status'
    ])
    
    # Data rows
    for row in report.rows:
        writer.writerow([
            row.invoice_number,
            row.invoice_date.isoformat(),
            row.customer_name,
            row.customer_gstin or '',
            row.place_of_supply,
            f"{row.taxable_value:.2f}",
            f"{row.cgst:.2f}",
            f"{row.sgst:.2f}",
            f"{row.igst:.2f}",
            f"{row.total_gst:.2f}",
            f"{row.grand_total:.2f}",
            row.status
        ])
    
    # Summary row
    writer.writerow([])
    writer.writerow(['TOTALS', '', '', '', '',
                     f"{report.summary.total_taxable_value:.2f}",
                     f"{report.summary.total_cgst:.2f}",
                     f"{report.summary.total_sgst:.2f}",
                     f"{report.summary.total_igst:.2f}",
                     f"{report.summary.total_gst:.2f}",
                     f"{report.summary.total_grand_total:.2f}",
                     f"{report.summary.count_invoices} invoices ({report.summary.count_cancelled} cancelled)"])
    
    # Return as downloadable file
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=sales_register_{from_date}_{to_date}.csv"
        }
    )


@router.get("/purchases/export/csv")
def export_purchase_register_csv(
    from_date: date = Query(...),
    to_date: date = Query(...),
    include_cancelled: bool = Query(False),
    supplier_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Purchase Register to CSV.
    
    CSV format matches on-screen report exactly.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    report = report_service.generate_purchase_register(
        db=db,
        from_date=from_date,
        to_date=to_date,
        include_cancelled=include_cancelled,
        supplier_id=supplier_id
    )
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Supplier Name', 'Supplier GSTIN', 'Supplier Invoice Number', 'Invoice Date',
        'Taxable Value', 'Input CGST', 'Input SGST', 'Input IGST',
        'Total Input GST', 'Grand Total', 'Status'
    ])
    
    # Data rows
    for row in report.rows:
        writer.writerow([
            row.supplier_name,
            row.supplier_gstin or '',
            row.supplier_invoice_number,
            row.invoice_date.isoformat(),
            f"{row.taxable_value:.2f}",
            f"{row.input_cgst:.2f}",
            f"{row.input_sgst:.2f}",
            f"{row.input_igst:.2f}",
            f"{row.total_input_gst:.2f}",
            f"{row.grand_total:.2f}",
            row.status
        ])
    
    # Summary row
    writer.writerow([])
    writer.writerow(['TOTALS', '', '', '',
                     f"{report.summary.total_taxable_value:.2f}",
                     f"{report.summary.total_input_cgst:.2f}",
                     f"{report.summary.total_input_sgst:.2f}",
                     f"{report.summary.total_input_igst:.2f}",
                     f"{report.summary.total_input_gst:.2f}",
                     f"{report.summary.total_grand_total:.2f}",
                     f"{report.summary.count_purchases} purchases ({report.summary.count_cancelled} cancelled)"])
    
    # Return as downloadable file
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=purchase_register_{from_date}_{to_date}.csv"
        }
    )

@router.get("/customers", response_model=CustomerReportResponse)
def get_customer_report(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    customer_id: Optional[int] = Query(None, description="Filter by customer ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate customer-wise sales summary.
    
    Shows sales value and GST collected per customer with B2B vs B2C split.
    Optional: Filter by specific customer_id.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_customer_report(
        db=db,
        from_date=from_date,
        to_date=to_date,
        customer_id=customer_id
    )


@router.get("/suppliers", response_model=SupplierReportResponse)
def get_supplier_report(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    supplier_id: Optional[int] = Query(None, description="Filter by supplier ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate supplier-wise purchase summary.
    
    Shows purchase value and input GST per supplier with Registered vs Unregistered split.
    Optional: Filter by specific supplier_id.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_supplier_report(
        db=db,
        from_date=from_date,
        to_date=to_date,
        supplier_id=supplier_id
    )


@router.get("/product-hsn", response_model=ProductHSNReportResponse)
def get_product_hsn_report(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    product_id: Optional[int] = Query(None, description="Filter by product ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate product/HSN-wise sales summary for GSTR-1 preparation.
    
    Aggregates quantity sold and GST collected by HSN code.
    Optional: Filter by specific product_id.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_product_hsn_report(
        db=db,
        from_date=from_date,
        to_date=to_date,
        product_id=product_id
    )


@router.get("/inventory", response_model=InventoryReportResponse)
def get_inventory_report(
    as_of_date: date = Query(..., description="Report date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate inventory report as of a specific date.
    
    Shows opening stock, purchases, sales, and closing stock per product.
    Note: No FIFO/LIFO valuation in Phase-1.
    """
    return report_service.generate_inventory_report(
        db=db,
        as_of_date=as_of_date
    )


@router.get("/business-ledger", response_model=BusinessSummaryLedger)
def get_business_ledger(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate Business Summary Ledger (informational only).
    
    Shows total sales, purchases, output GST, input GST, and net GST position.
    NOTE: NOT a statutory accounting ledger. For accounting, use proper accounting software.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_business_summary_ledger(
        db=db,
        from_date=from_date,
        to_date=to_date
    )


@router.get("/gstr-1-export", response_model=GSTRReadyResponse)
def get_gstr_ready_export(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Generate GSTR-1 ready export (preparation only).
    
    Exports B2B invoices, B2C summary, and HSN summary in GSTR-1 format.
    Manual filing on gst.gov.in required.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    return report_service.generate_gstr_ready_export(
        db=db,
        from_date=from_date,
        to_date=to_date
    )


# ========== EXCEL EXPORT ENDPOINTS ==========

@router.get("/sales/export/xlsx")
def export_sales_register_xlsx(
    from_date: date = Query(...),
    to_date: date = Query(...),
    include_cancelled: bool = Query(False),
    customer_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Sales Register to Excel (XLSX format).
    
    Creates a formatted spreadsheet with data and summary.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    report = report_service.generate_sales_register(
        db=db,
        from_date=from_date,
        to_date=to_date,
        include_cancelled=include_cancelled,
        customer_id=customer_id
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Register"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Invoice No', 'Date', 'Customer', 'GSTIN', 'Place of Supply',
               'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total GST', 'Grand Total', 'Status']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, row in enumerate(report.rows, 2):
        ws.append([
            row.invoice_number,
            row.invoice_date.isoformat(),
            row.customer_name,
            row.customer_gstin or '',
            row.place_of_supply,
            row.taxable_value,
            row.cgst,
            row.sgst,
            row.igst,
            row.total_gst,
            row.grand_total,
            row.status
        ])
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num >= 6:  # Number columns
                cell.alignment = Alignment(horizontal='right')
    
    # Summary
    summary_row = len(report.rows) + 3
    ws.cell(row=summary_row, column=1).value = "TOTALS"
    ws.cell(row=summary_row, column=6).value = report.summary.total_taxable_value
    ws.cell(row=summary_row, column=7).value = report.summary.total_cgst
    ws.cell(row=summary_row, column=8).value = report.summary.total_sgst
    ws.cell(row=summary_row, column=9).value = report.summary.total_igst
    ws.cell(row=summary_row, column=10).value = report.summary.total_gst
    ws.cell(row=summary_row, column=11).value = report.summary.total_grand_total
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col_num)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
        if col_num >= 6:
            cell.alignment = Alignment(horizontal='right')
    
    # Summary text
    ws.cell(row=summary_row + 1, column=1).value = f"Total Invoices: {report.summary.count_invoices} (Cancelled: {report.summary.count_cancelled})"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 14
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=sales_register_{from_date}_{to_date}.xlsx"
        }
    )


@router.get("/purchases/export/xlsx")
def export_purchase_register_xlsx(
    from_date: date = Query(...),
    to_date: date = Query(...),
    include_cancelled: bool = Query(False),
    supplier_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Purchase Register to Excel (XLSX format).
    
    Creates a formatted spreadsheet with data and summary.
    """
    if to_date < from_date:
        raise HTTPException(status_code=400, detail="to_date must be >= from_date")
    
    report = report_service.generate_purchase_register(
        db=db,
        from_date=from_date,
        to_date=to_date,
        include_cancelled=include_cancelled,
        supplier_id=supplier_id
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Register"
    
    # Styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    summary_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Supplier', 'GSTIN', 'Supplier Invoice No', 'Invoice Date',
               'Taxable Value', 'Input CGST', 'Input SGST', 'Input IGST',
               'Total Input GST', 'Grand Total', 'Status']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, row in enumerate(report.rows, 2):
        ws.append([
            row.supplier_name,
            row.supplier_gstin or '',
            row.supplier_invoice_number,
            row.invoice_date.isoformat(),
            row.taxable_value,
            row.input_cgst,
            row.input_sgst,
            row.input_igst,
            row.total_input_gst,
            row.grand_total,
            row.status
        ])
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num >= 5:  # Number columns
                cell.alignment = Alignment(horizontal='right')
    
    # Summary
    summary_row = len(report.rows) + 3
    ws.cell(row=summary_row, column=1).value = "TOTALS"
    ws.cell(row=summary_row, column=5).value = report.summary.total_taxable_value
    ws.cell(row=summary_row, column=6).value = report.summary.total_input_cgst
    ws.cell(row=summary_row, column=7).value = report.summary.total_input_sgst
    ws.cell(row=summary_row, column=8).value = report.summary.total_input_igst
    ws.cell(row=summary_row, column=9).value = report.summary.total_input_gst
    ws.cell(row=summary_row, column=10).value = report.summary.total_purchase_value
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col_num)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
        if col_num >= 5:
            cell.alignment = Alignment(horizontal='right')
    
    # Summary text
    ws.cell(row=summary_row + 1, column=1).value = f"Total Purchases: {report.summary.count_purchases} (Cancelled: {report.summary.count_cancelled} if any)"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 14
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=purchase_register_{from_date}_{to_date}.xlsx"
        }
    )